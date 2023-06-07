import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np




def save_obj(obj, name ):
    f= open(name + '.pkl', 'wb')
    pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)

def load_obj(name):
    with open(name + '.pkl', 'rb') as f:
        return pickle.load(f)

def writeListXls(fileName, header, values):
    wb = Workbook()
    ws1 = wb.active

    ws1.cell(column=1, row=1, value=header)

    row = 2
    for g in values:
        ws1.cell(column=1, row=row, value=g)
        row += 1


    wb.save(filename = fileName)

def writeMatrixXls(fileName, header, values):
    wb = Workbook()
    ws1 = wb.active

    i=1
    for h in header:
        ws1.cell(column=i, row=1, value=h)
        i+=1

    c=1
    for g in values:
        row = 2
        for a in g:
            ws1.cell(column=c, row=row, value=a)
            row += 1
        c+=1


    wb.save(filename = fileName)






def computeMSE(x_true, x_res):
    actual, pred = np.array(x_true), np.array(x_res)
    return np.square(np.subtract(actual,pred)).mean()

def writeArray(namefile, list_of_array):
    # save array into csv file
    list_of_array=np.array(list_of_array)
    np.savetxt(namefile+".csv", list_of_array,
                  delimiter=",")

def readArray(namefile):
    return np.loadtxt(namefile+".csv",delimiter=",")

def writeFile(nameFile, string_):
    text_file = open(nameFile, "w")
    text_file.write(string_)
    text_file.close()



def writeListResultsXls(fileName, header, results):
    wb = Workbook()
    ws1 = wb.active

    c=1
    for h in header:
        ws1.cell(column=c, row=1, value=str(h))
        c+=1

    c=1
    for g in results:
        row = 2
        for e in g:
            ws1.cell(column=c, row=row, value=e)
            row += 1
        c+=1

    wb.save(filename=fileName)

def readXls(filename, sheet, col_name, flagFloat=True):
    wb = load_workbook(filename=filename)
    sheet = wb[sheet]
    fitness_col=sheet[col_name]
    l=[]
    for i in range(1,len(fitness_col)):
        print(i)
        cell=fitness_col[i]
        v=cell.value
        print(v)
        if flagFloat==True:
            v=float(v)
        l.append(v)
    return l


